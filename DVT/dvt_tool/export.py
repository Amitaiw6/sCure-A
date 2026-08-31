"""Campaign exports (SRS-DVT-093, 110): JSON snapshot, per-test CSV, one
XLSX workbook (Summary + one sheet per test), Markdown report. Written to a
local folder; the sync layer pushes that folder to Google Drive.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .catalog import Catalog
from .store import Store

VERDICT_FILL = {"PASS": "C6EFCE", "FAIL": "FFC7CE", "BLOCKED": "FFEB9C", "WAIVED": "D9D2E9"}


def _stamp() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


class Exporter:
    def __init__(self, catalog: Catalog, store: Store, out_dir: str | Path, campaign: str = "sCure-DVT"):
        self.cat, self.store, self.out, self.campaign = catalog, store, Path(out_dir), campaign
        self.out.mkdir(parents=True, exist_ok=True)

    # ---------------- all ----------------
    def export_all(self) -> list[Path]:
        files = [self.export_json(), self.export_xlsx(), self.export_report_md()]
        files += self.export_csvs()
        return files

    # ---------------- JSON snapshot ----------------
    def snapshot(self) -> dict:
        runs = self.store.runs(include_supplementary=True)
        for r in runs:
            r["values"] = self.store.values(r["run_id"])
            r["redlines"] = self.store.redlines(r["run_id"])
        return {"campaign": self.campaign, "exportedAt": _stamp(), "catalogVersion": self.cat.version,
                "units": self.store.units(), "runs": runs, "ncrs": self.store.ncrs(), "waivers": self.store.waivers(),
                "attachments": self.store.attachments(), "events": self.store.events(limit=5000)}

    def export_json(self) -> Path:
        p = self.out / f"{self.campaign}.campaign.json"
        p.write_text(json.dumps(self.snapshot(), indent=2, ensure_ascii=False), encoding="utf-8")
        return p

    # ---------------- CSV per test ----------------
    def _test_rows(self, test_id: str) -> tuple[list[str], list[list]]:
        fields = [f["name"] for f in self.cat.tests[test_id].get("data_fields") or []]
        variant_keys = sorted({k for r in self.store.runs(test_id, include_supplementary=True) for k in r["variant"]})
        header = ["run_id", "unit", *variant_keys, "repetition", "status", "verdict", "verdict_detail", "operator",
                  "witness", "started_at", "finished_at", *fields, "redlines", "reject_reason"]
        rows = []
        for r in self.store.runs(test_id, include_supplementary=True):
            v = self.store.values(r["run_id"])
            rows.append([r["run_id"], r["unit_id"], *[r["variant"].get(k) for k in variant_keys], r["repetition"],
                         r["status"], r["verdict"], r["verdict_detail"], r["operator"], r["witness"], r["started_at"],
                         r["finished_at"], *[v.get(f) for f in fields], len(self.store.redlines(r["run_id"])), r["reject_reason"]])
        return header, rows

    def export_csvs(self) -> list[Path]:
        d = self.out / "csv"; d.mkdir(exist_ok=True)
        out = []
        for tid in self.cat.ordered_test_ids():
            header, rows = self._test_rows(tid)
            p = d / f"{tid}.csv"
            with p.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f); w.writerow(header); w.writerows(rows)
            out.append(p)
        return out

    # ---------------- XLSX ----------------
    def export_xlsx(self) -> Path:
        wb = Workbook()
        ws = wb.active; ws.title = "Summary"
        bold = Font(bold=True)
        ws.append([f"{self.campaign} — DVT campaign", "", f"catalog v{self.cat.version}", f"exported {_stamp()}"])
        ws["A1"].font = bold
        ws.append([])
        ws.append(["Test", "Title", "Phase", "Applicability", "Runs", "Done", "PASS", "FAIL", "BLOCKED", "WAIVED", "Rolled-up"])
        for c in ws[3]: c.font = bold
        from .engine import Engine   # local import: engine imports store/catalog only
        eng = Engine(self.cat, self.store)
        for tid in self.cat.ordered_test_ids():
            t = self.cat.tests[tid]
            runs = [r for r in self.store.runs(tid) if r["status"] != "REJECTED"]
            cnt = lambda v: sum(1 for r in runs if r["verdict"] == v)  # noqa: E731
            roll = eng.test_verdict(tid) or ""
            ws.append([tid, t["title"], self.cat.phase_of(tid)["id"], t["applicability"]["rule"], len(runs),
                       sum(1 for r in runs if r["status"] == "DONE"), cnt("PASS"), cnt("FAIL"), cnt("BLOCKED"), cnt("WAIVED"), roll])
            if roll in VERDICT_FILL:
                ws.cell(row=ws.max_row, column=11).fill = PatternFill("solid", fgColor=VERDICT_FILL[roll])
        ws.append([])
        ws.append(["Open NCRs", len(self.store.ncrs(open_only=True))])
        ws.append(["Waivers", len(self.store.waivers())])
        for i, w in enumerate([12, 60, 7, 13, 6, 6, 6, 6, 9, 8, 11], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for tid in self.cat.ordered_test_ids():
            header, rows = self._test_rows(tid)
            s = wb.create_sheet(tid[-7:])            # sheet names max 31 chars
            s.append(header)
            for c in s[1]: c.font = bold
            for row in rows:
                s.append([json.dumps(x) if isinstance(x, (dict, list)) else x for x in row])
                v = row[header.index("verdict")]
                if v in VERDICT_FILL:
                    s.cell(row=s.max_row, column=header.index("verdict") + 1).fill = PatternFill("solid", fgColor=VERDICT_FILL[v])
            s.freeze_panes = "C2"
        n = wb.create_sheet("NCR")
        n.append(["NCR", "Run", "Opened", "By", "Description", "Disposition", "Closed", "Closed by"])
        for c in n[1]: c.font = bold
        for r in self.store.ncrs():
            n.append([r["ncr_id"], r["run_id"], r["opened_at"], r["opened_by"], r["description"], r["disposition"], r["closed_at"], r["closed_by"]])
        p = self.out / f"{self.campaign}.xlsx"
        wb.save(p)
        return p

    # ---------------- Markdown report ----------------
    def export_report_md(self) -> Path:
        from .engine import Engine
        eng = Engine(self.cat, self.store)
        prog = eng.progress()
        lines = [f"# {self.campaign} — DVT Verification Report (working)", "",
                 f"Catalog v{self.cat.version} · exported {_stamp()}", "",
                 f"**Runs:** {prog['done']}/{prog['total']} done · PASS {prog['PASS']} · FAIL {prog['FAIL']} · "
                 f"BLOCKED {prog['BLOCKED']} · WAIVED {prog['WAIVED']} · rejected {prog['rejected']} · open NCRs {prog['openNcrs']}", "",
                 "## Per unit", "", "| Unit | Phase | Done / owed |", "|---|---|---|"]
        for u, d in prog["perUnit"].items():
            lines.append(f"| {u} | {d['phase']} | {d['done']} / {d['total']} |")
        lines += ["", "## Per test", "", "| Test | Title | Rolled-up | PASS | FAIL | BLOCKED | WAIVED | Pending |", "|---|---|---|---|---|---|---|---|"]
        for tid in self.cat.ordered_test_ids():
            runs = [r for r in self.store.runs(tid) if r["status"] != "REJECTED"]
            c = lambda v: sum(1 for r in runs if r["verdict"] == v)  # noqa: E731
            pending = sum(1 for r in runs if r["status"] != "DONE")
            lines.append(f"| {tid} | {self.cat.tests[tid]['title']} | {eng.test_verdict(tid) or '—'} | {c('PASS')} | {c('FAIL')} | {c('BLOCKED')} | {c('WAIVED')} | {pending} |")
        ncrs = self.store.ncrs()
        if ncrs:
            lines += ["", "## Non-conformances", "", "| NCR | Run | Description | Disposition |", "|---|---|---|---|"]
            for r in ncrs:
                lines.append(f"| {r['ncr_id']} | {r['run_id']} | {r['description']} | {r['disposition'] or 'open'} |")
        w = self.store.waivers()
        if w:
            lines += ["", "## Waivers", "", "| Run | Approver | Rationale |", "|---|---|---|"]
            for r in w:
                lines.append(f"| {r['run_id']} | {r['approver']} | {r['rationale']} |")
        # as-run procedure: every redline, per run (SRS-DVT-086 / NASA item 8)
        redlined = [(r, self.store.redlines(r["run_id"])) for r in self.store.runs(include_supplementary=True)]
        redlined = [(r, rl) for r, rl in redlined if rl]
        if redlined:
            lines += ["", "## As-run deviations (redlines)", "", "| Run | Step | Performed as | Reason | By | At |", "|---|---|---|---|---|---|"]
            for r, rls in redlined:
                for x in rls:
                    lines.append(f"| {r['run_id']} | {x['step_index'] + 1} | {x['as_run']} | {x['reason']} | {x['by_whom']} | {x['at']} |")
        unverified = [tid for tid in self.cat.ordered_test_ids() if eng.test_verdict(tid) not in ("PASS", "WAIVED")]
        lines += ["", "## Not yet verified (SRS-DVT-053)", ""] + [f"- {t}" for t in unverified] if unverified else ["", "All tests verified."]
        p = self.out / f"{self.campaign}.report.md"
        p.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return p
