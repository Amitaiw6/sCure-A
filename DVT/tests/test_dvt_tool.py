import sys
from datetime import date, timedelta
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from dvt_tool.catalog import Catalog  # noqa: E402
from dvt_tool import criteria  # noqa: E402
from dvt_tool.store import Store  # noqa: E402
from dvt_tool.engine import Engine  # noqa: E402
from dvt_tool.export import Exporter  # noqa: E402
from dvt_tool.drive import SyncConfig, Syncer, FolderBackend  # noqa: E402

CATALOG = ROOT / "catalog" / "DVT_test_catalog.yaml"


@pytest.fixture
def cat():
    return Catalog.load(CATALOG)


@pytest.fixture
def world(tmp_path, cat):
    store = Store(tmp_path / "campaign.db")
    eng = Engine(cat, store)
    return cat, store, eng, tmp_path


# ---------------- catalog ----------------
def test_catalog_loads_and_run_set_matches_srs(cat):
    assert len(cat.tests) == 28
    runs = cat.runs()
    # SRS Rev B: 356 applicable runs; ENV-001 has unit=null (CONFIRM) so it contributes 0 until decided
    assert len(runs) == 355
    assert len(cat.runs("DVT-THM-001")) == 50          # 10 matrix rows x 5 units
    assert len(cat.runs("DVT-SAF-008")) == 40          # 8 reps x 5 units
    assert len(cat.runs("DVT-SAF-018")) == 7           # 7 cases, single unit
    assert len(cat.runs("DVT-SAF-001")) == 1
    r = cat.runs("DVT-THM-001")[0]
    assert r.run_id.startswith("DVT-THM-001|UUT-01|setpoint=30,mains_voltage=230|1")
    assert cat.parent_row_for("DVT-THM-003", {"cooling_mode": "normal"}) == {"test_id": "DVT-THM-001", "setpoint": 80, "mains_voltage": 110}


def test_catalog_validation_catches_bad_entries(tmp_path, cat):
    import yaml
    data = yaml.safe_load(CATALOG.read_text(encoding="utf-8"))
    data["tests"][0]["procedure_steps"][2]["capture"].append("no_such_field")
    p = tmp_path / "bad.yaml"; p.write_text(yaml.safe_dump(data), encoding="utf-8")
    with pytest.raises(Exception, match="unknown field"):
        Catalog.load(p)


# ---------------- criteria ----------------
def test_criteria_pass_fail_blocked(cat):
    th = cat.thresholds
    expr = cat.tests["DVT-THM-002"]["pass_criteria"]
    ok = dict(max_temp_right=70, max_temp_left=72, max_temp_back=74.9, max_temp_door=60, protective_trip_occurred=False, final_dtdt=0.1, tj_margin=16)
    assert criteria.evaluate(expr, ok, th) == ("PASS", "")
    assert criteria.evaluate(expr, {**ok, "max_temp_back": 75.1}, th)[0] == "FAIL"
    v, d = criteria.evaluate(expr, {k: v for k, v in ok.items() if k != "tj_margin"}, th)
    assert v == "BLOCKED" and "tj_margin" in d
    # SAF-019 references a threshold that is still null -> BLOCKED, never PASS
    expr19 = cat.tests["DVT-SAF-019"]["pass_criteria"]
    v, d = criteria.evaluate(expr19, dict(alarm_raised=True, alarm_identifies_correct_side=True, message_names_specific_fan=False,
                                          claim_is_correct=True, peak_temp_after_shutdown=70, post_shutdown_overshoot=3,
                                          false_alarm_on_other_sides=False, recovery_requires_ack=True), th)
    assert v == "BLOCKED" and "protective_shutdown" in d
    # string / enum / OR / abs
    assert criteria.evaluate("which == 'a' AND (x == false OR y == true) AND abs(e) <= 1", dict(which="a", x=True, y=True, e=-0.5)) == ("PASS", "")
    with pytest.raises(criteria.CriteriaError):
        criteria.evaluate("__import__('os')", {})


# ---------------- engine: gates, next action, verdict, rollup ----------------
def _cal_all(cat, store):
    until = (date.today() + timedelta(days=365)).isoformat()
    for t in cat.tests.values():
        for e in t.get("equipment") or []:
            store.set_calibration(e["name"], "CAL-1", until)


def test_wizard_gates_and_next_action(world):
    cat, store, eng, _ = world
    na = eng.next_action("UUT-02")
    assert na.phase["id"] == 0 and na.run is None                        # config not frozen
    store.freeze_config("UUT-02", "amitai", "SC000002")
    na = eng.next_action("UUT-02")
    assert na.run["test_id"] == "DVT-ELE-001"                           # earth first
    codes = {b.code for b in na.blockers}
    assert "TRR" in codes and "CAL" in codes and "PHASE" not in codes
    store.sign_phase("UUT-02", 1, "amitai", ["ok"])
    _cal_all(cat, store)
    na = eng.next_action("UUT-02")
    assert na.blockers == []
    # a later-phase run is blocked by the unclosed phase and by the earth rule
    saf = store.runs("DVT-SAF-012", "UUT-02")[0]
    codes = {b.code for b in eng.blockers_for(saf)}
    assert {"PHASE", "EARTH", "TRR"} <= codes


def test_finish_pass_fail_ncr_waiver_and_rollup(world):
    cat, store, eng, _ = world
    store.freeze_config("UUT-01", "a"); store.sign_phase("UUT-01", 1, "a", []); _cal_all(cat, store)
    run = store.runs("DVT-ELE-001", "UUT-01")[0]
    store.start_run(run["run_id"], "a"); store.confirm_preconditions(run["run_id"], "a")
    store.set_values(run["run_id"], dict(max_resistance=0.05, pe_uninterrupted=True), "a")
    assert eng.finish(run["run_id"], "a") == ("PASS", "")
    # ELE-002 fails -> NCR opened, rollup FAIL, then waived -> WAIVED counted separately
    r2 = store.runs("DVT-ELE-002", "UUT-01")[0]
    store.start_run(r2["run_id"], "a"); store.set_values(r2["run_id"], dict(insulation_resistance_lnpe=4), "a")
    v, _ = eng.finish(r2["run_id"], "a")
    assert v == "FAIL" and len(store.ncrs(open_only=True)) == 1
    assert eng.test_verdict("DVT-ELE-002", "UUT-01") == "FAIL"
    with pytest.raises(ValueError):
        store.waive(r2["run_id"], "", "no approver", "a")
    store.waive(r2["run_id"], "QA lead", "engineering limit, not a 62368 requirement", "a")
    assert eng.test_verdict("DVT-ELE-002", "UUT-01") == "WAIVED"
    p = eng.progress()
    assert p["PASS"] == 1 and p["WAIVED"] == 1 and p["FAIL"] == 0
    # resume after restart: in-progress run is the next action (SRS-DVT-088)
    r3 = store.runs("DVT-ELE-003", "UUT-01")[0]
    store.start_run(r3["run_id"], "a"); store.set_step(r3["run_id"], 2)
    eng2 = Engine(cat, Store(store.path))
    assert eng2.next_action("UUT-01").run["run_id"] == r3["run_id"]


def test_sweep_rollup_and_reject_rescinds_children(world):
    cat, store, eng, _ = world
    rows = store.runs("DVT-THM-001", "UUT-03")
    good = dict(uniformity_pp=3, overshoot=1, steady_state_error=0.5, limit_cycle_amplitude=0.5, settling_time=20)
    for r in rows:
        store.start_run(r["run_id"], "a"); store.set_values(r["run_id"], good, "a"); eng.finish(r["run_id"], "a")
    assert eng.test_verdict("DVT-THM-001", "UUT-03") == "PASS"
    bad = rows[5]
    store.set_values(bad["run_id"], {**good, "overshoot": 4}, "a"); eng.finish(bad["run_id"], "a")
    assert eng.test_verdict("DVT-THM-001", "UUT-03") == "FAIL"          # one row fails -> rolled-up FAIL (SRS-DVT-096)
    # co-executed child linked to the parent row; rejecting the parent rescinds the child (SRS-DVT-099)
    child = [r for r in store.runs("DVT-THM-003", "UUT-03") if r["variant"] == {"cooling_mode": "fast"}][0]
    parent = [r for r in rows if r["variant"] == {"setpoint": 80, "mains_voltage": 230}][0]
    store.link_parent(child["run_id"], parent["run_id"])
    affected = store.reject_run(parent["run_id"], "ambient left the band", "a")
    assert set(affected) == {parent["run_id"], child["run_id"]}
    assert store.run(child["run_id"])["status"] == "REJECTED" and "parent rejected" in store.run(child["run_id"])["reject_reason"]
    assert eng.test_verdict("DVT-THM-001", "UUT-03") == "FAIL"          # the other failing row still counts


# ---------------- export + sync ----------------
def test_export_and_folder_sync(world):
    cat, store, eng, tmp = world
    run = store.runs("DVT-ELE-001", "UUT-01")[0]
    store.start_run(run["run_id"], "a"); store.set_values(run["run_id"], dict(max_resistance=0.2, pe_uninterrupted=True), "a")
    eng.finish(run["run_id"], "a")
    store.add_redline(run["run_id"], 1, "measured with 10 A instead of 25 A", "tester limit", "a")
    exp = Exporter(cat, store, tmp / "export", "TestCampaign")
    files = exp.export_all()
    names = {f.name for f in files}
    assert {"TestCampaign.campaign.json", "TestCampaign.xlsx", "TestCampaign.report.md", "DVT-ELE-001.csv"} <= names
    csv_text = (tmp / "export" / "csv" / "DVT-ELE-001.csv").read_text()
    assert "FAIL" in csv_text and "0.2" in csv_text
    from openpyxl import load_workbook
    wb = load_workbook(tmp / "export" / "TestCampaign.xlsx")
    assert "Summary" in wb.sheetnames and "ELE-001" in wb.sheetnames and "NCR" in wb.sheetnames
    # folder backend: everything lands under the synced folder, queue drains
    drive = tmp / "GoogleDrive" / "sCure DVT"
    syncer = Syncer(SyncConfig(mode="folder", folder_path=drive, campaign="TestCampaign"), store, tmp / "export")
    st = syncer.sync(files)
    assert st.ok and (drive / "TestCampaign.xlsx").exists() and (drive / "csv" / "DVT-ELE-001.csv").exists()
    assert store.pending_sync() == []
    # a failing backend never loses data: queue stays, error visible
    class Broken(FolderBackend):
        def push(self, files, root): raise OSError("network down")
    store.queue("run", run["run_id"])
    bad = Syncer(SyncConfig(mode="folder", folder_path=drive), store, tmp / "export", backend=Broken(drive))
    st = bad.sync(files)
    assert not st.ok and st.pending == 1 and "network down" in st.last_error
