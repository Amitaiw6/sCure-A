#!/usr/bin/env python3
"""
uv_led_experiment.py - LED temperature vs UV intensity characterization.

Steps the UV output from --start to --stop in --step increments. At each
intensity it waits for the LED-panel temperature (TC-08 average across the
connected thermocouples) to stabilize, records the settled average, then moves
on. Ends with UV OFF and writes:

  led_temp_vs_uv.xlsx - Results sheet (intensity vs settled temp + chart)
                        and a Samples sheet with the full time series
  led_temp_vs_uv.png  - the same curve as an image

Stdout doubles as the monitoring log (uv_experiment_viewer.py tails it):
  "  HH:MM:SS  uv  35  temp  47.23"        every sample
  "STEP 35% -> 47.1 C (stable after 3.2 min)"  every completed step

Usage:
    python uv_led_experiment.py [--start 10] [--stop 100] [--step 5]
        [--wavelength 405] [--stable-band 0.5] [--stable-minutes 2]
        [--max-step-minutes 12] [--out-dir .]
"""

import argparse
import datetime as dt
import json
import os
import sys
import time
import urllib.request

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from calibrate_scure_temp import TC08  # noqa: E402

SURFACE, SERIES, INK, MUTED, GRID, BASELINE = (
    "#fcfcfb", "#2a78d6", "#0b0b0b", "#898781", "#e1e0d9", "#c3c2b7")


def api(host, path, method="GET", timeout=8):
    req = urllib.request.Request(f"http://{host}:3001/api{path}", method=method)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.load(r)


def set_uv(host, on, intensity=None, wavelength=405):
    q = f"?on={1 if on else 0}"
    if on:
        q += f"&intensity={intensity}&wavelength={wavelength}"
    resp = api(host, "/uv" + q, method="POST")
    if not resp.get("ok"):
        raise RuntimeError(f"UV command failed: {resp.get('message')}")
    return resp


def wait_stable(tc, band, stable_sec, max_sec, label):
    """Sample every 5 s until the average temp holds within `band` degC over
    `stable_sec`. Returns (settled_mean, samples[(datetime, temp)])."""
    window, samples = [], []
    t0 = time.monotonic()
    while True:
        now = time.monotonic()
        avg = tc.average()
        if avg is not None:
            stamp = dt.datetime.now()
            samples.append((stamp, avg))
            window.append((now, avg))
            window = [(t, x) for t, x in window if now - t <= stable_sec]
            print(f"  {stamp:%H:%M:%S}  uv  {label}  temp  {avg:.2f}", flush=True)
            covered = window and now - window[0][0] >= stable_sec * 0.9
            if covered and now - t0 >= stable_sec * 1.2:
                span = max(x for _, x in window) - min(x for _, x in window)
                if span <= band:
                    tail = [x for _, x in window][-12:]
                    return sum(tail) / len(tail), samples
        if now - t0 >= max_sec:
            tail = [x for _, x in window][-12:] or [avg or 0.0]
            print(f"  step {label}%: max time reached, using last minute mean", flush=True)
            return sum(tail) / len(tail), samples
        time.sleep(5)


def write_outputs(results, all_samples, out_dir, wavelength):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.marker import Marker
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "LED temp vs UV"
    ws.append([f"UV intensity % ({wavelength} nm)", "LED temp (degC, TC-08 avg)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for pct, temp in results:
        ws.append([pct, round(temp, 2)])
    chart = LineChart()
    chart.title = f"LED temperature vs UV intensity ({wavelength} nm)"
    chart.x_axis.title = "UV intensity (%)"
    chart.y_axis.title = "degC"
    chart.height, chart.width = 10, 18
    n = len(results)
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=n + 1),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = SERIES.lstrip("#")
    s.graphicalProperties.line.width = 25000
    s.marker = Marker(symbol="circle", size=6)
    ws.add_chart(chart, "D2")

    ws2 = wb.create_sheet("Samples")
    ws2.append(["Time", "UV intensity %", "LED temp (degC)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for stamp, pct, temp in all_samples:
        ws2.append([stamp.strftime("%Y-%m-%d %H:%M:%S"), pct, round(temp, 2)])
    ws2.column_dimensions["A"].width = 20

    xlsx = os.path.join(out_dir, "led_temp_vs_uv.xlsx")
    wb.save(xlsx)

    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    xs = [p for p, _ in results]
    ys = [t for _, t in results]
    fig, ax = plt.subplots(figsize=(9, 5.5), facecolor=SURFACE)
    ax.set_facecolor(SURFACE)
    ax.plot(xs, ys, color=SERIES, linewidth=2, marker="o", markersize=5)
    ax.set_title(f"LED temperature vs UV intensity ({wavelength} nm)",
                 color=INK, fontsize=12, loc="left", pad=12)
    ax.set_xlabel("UV intensity (%)", color=MUTED)
    ax.set_ylabel("\N{DEGREE SIGN}C", color=MUTED)
    ax.grid(color=GRID, linewidth=0.8)
    ax.tick_params(colors=MUTED)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color(BASELINE)
    fig.tight_layout()
    png = os.path.join(out_dir, "led_temp_vs_uv.png")
    fig.savefig(png, dpi=150, facecolor=SURFACE)
    plt.close(fig)
    return xlsx, png


def main():
    ap = argparse.ArgumentParser(description="LED temp vs UV intensity sweep.")
    ap.add_argument("--host", default="192.168.154.141")
    ap.add_argument("--start", type=int, default=10)
    ap.add_argument("--stop", type=int, default=100)
    ap.add_argument("--step", type=int, default=5)
    ap.add_argument("--wavelength", type=int, default=405, choices=(405, 450))
    ap.add_argument("--stable-band", type=float, default=0.5,
                    help="max degC span over the window to count as settled")
    ap.add_argument("--stable-minutes", type=float, default=2.0)
    ap.add_argument("--max-step-minutes", type=float, default=12.0)
    ap.add_argument("--out-dir", default=".")
    args = ap.parse_args()

    state = api(args.host, "/state")
    if state.get("doorClosed") is not True:
        raise SystemExit("Door is not confirmed closed - close it and rerun (UV interlock).")

    os.makedirs(args.out_dir, exist_ok=True)
    intensities = list(range(args.start, args.stop + 1, args.step))
    print(f"UV sweep {args.wavelength} nm: {intensities[0]}-{intensities[-1]}% "
          f"in {args.step}% steps ({len(intensities)} points). "
          f"Settled = span<{args.stable_band} C over {args.stable_minutes:g} min.",
          flush=True)

    tc = TC08()
    results, all_samples = [], []
    try:
        for pct in intensities:
            set_uv(args.host, True, pct, args.wavelength)
            print(f"UV -> {pct}%", flush=True)
            settled, samples = wait_stable(
                tc, args.stable_band, args.stable_minutes * 60,
                args.max_step_minutes * 60, pct)
            results.append((pct, settled))
            all_samples.extend((t, pct, x) for t, x in samples)
            mins = len(samples) * 5 / 60
            print(f"STEP {pct}% -> {settled:.1f} C (settled after {mins:.1f} min)",
                  flush=True)
    finally:
        try:
            set_uv(args.host, False)
            print("UV -> OFF", flush=True)
        except Exception as e:      # noqa: BLE001
            print(f"WARNING: could not turn UV off: {e} - turn it off manually!",
                  flush=True)
        tc.close()

    if not results:
        raise SystemExit("No results captured.")
    xlsx, png = write_outputs(results, all_samples, args.out_dir, args.wavelength)
    print(f"DONE: {len(results)} points")
    print(f"Excel: {xlsx}")
    print(f"Graph: {png}")


if __name__ == "__main__":
    main()
