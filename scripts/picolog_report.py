#!/usr/bin/env python3
"""
picolog_report.py - average-temperature report from a Pico TC-08 USB logger.

Two ways to get the data:

  live  - sample the TC-08 directly over USB (needs PicoLog 6 or PicoSDK
          installed so Windows has the usbtc08 driver + DLL):
            python scripts/picolog_report.py live --minutes 10 --interval 2
            python scripts/picolog_report.py live --channels 1,2,3 --tc-type K

  csv   - use a CSV exported from the PicoLog 6 app (File > Export as CSV):
            python scripts/picolog_report.py csv "C:\\path\\to\\export.csv"

Both modes produce, in --out-dir (default: current folder):
  <prefix>.png   - graph of the average temperature over time
  <prefix>.xlsx  - Excel with all readings, per-channel stats, the overall
                   average, and a native Excel line chart of the average
"""

import argparse
import csv
import datetime as dt
import math
import os
import sys
import time

# ---------------------------------------------------------------- data model

class Recording:
    """Times (datetime), channel names, and rows of per-channel temps (float or None)."""

    def __init__(self, channel_names):
        self.channel_names = list(channel_names)
        self.times = []
        self.rows = []

    def add(self, when, temps):
        self.times.append(when)
        self.rows.append([t if _is_num(t) else None for t in temps])

    def row_averages(self):
        """Average across channels for each sample (None if no channel read)."""
        out = []
        for row in self.rows:
            vals = [t for t in row if t is not None]
            out.append(sum(vals) / len(vals) if vals else None)
        return out

    def channel_stats(self):
        """Per-channel (name, mean, min, max, count) ignoring missing samples."""
        stats = []
        for i, name in enumerate(self.channel_names):
            vals = [row[i] for row in self.rows if row[i] is not None]
            if vals:
                stats.append((name, sum(vals) / len(vals), min(vals), max(vals), len(vals)))
            else:
                stats.append((name, None, None, None, 0))
        return stats

    def overall_average(self):
        vals = [a for a in self.row_averages() if a is not None]
        return sum(vals) / len(vals) if vals else None


def _is_num(v):
    return isinstance(v, (int, float)) and math.isfinite(v)


# ---------------------------------------------------------------- live (TC-08)

def _add_pico_dll_dirs():
    """Help picosdk find usbtc08.dll (PicoLog 6 keeps it outside the SDK path)."""
    for d in (r"C:\Program Files\Pico Technology\PicoLog",
              r"C:\Program Files\Pico Technology\SDK\lib"):
        if os.path.isdir(d):
            os.add_dll_directory(d)
            os.environ["PATH"] = d + os.pathsep + os.environ.get("PATH", "")


def capture_live(minutes, interval, channels, tc_type):
    import ctypes
    _add_pico_dll_dirs()
    try:
        from picosdk.usbtc08 import usbtc08 as tc08
    except OSError as e:
        raise SystemExit(
            "Could not load the usbtc08 driver DLL.\n"
            "Install PicoLog 6 or PicoSDK from https://www.picotech.com/downloads\n"
            f"(details: {e})"
        )

    handle = tc08.usb_tc08_open_unit()
    if handle == 0:
        raise SystemExit("No TC-08 found. Is it plugged in, and is another app (PicoLog) using it?")
    if handle < 0:
        raise SystemExit("TC-08 open failed - unplug/replug the USB and try again.")

    rec = Recording([f"CH{c}" for c in channels])
    try:
        tc08.usb_tc08_set_mains(handle, 50)
        # channel 0 is the cold junction and must stay enabled for compensation
        tc08.usb_tc08_set_channel(handle, 0, ord('C'))
        for c in range(1, 9):
            tc08.usb_tc08_set_channel(handle, c, ord(tc_type) if c in channels else 0)

        temp = (ctypes.c_float * 9)()
        overflow = ctypes.c_int16(0)
        units = tc08.USBTC08_UNITS["USBTC08_UNITS_CENTIGRADE"]

        end = time.monotonic() + minutes * 60
        n = 0
        print(f"Sampling {len(channels)} channel(s) every {interval:g}s for {minutes:g} min - Ctrl+C to stop early.")
        while time.monotonic() < end:
            ok = tc08.usb_tc08_get_single(handle, ctypes.byref(temp),
                                          ctypes.byref(overflow), units)
            now = dt.datetime.now()
            if ok:
                rec.add(now, [float(temp[c]) for c in channels])
                n += 1
                avg = rec.row_averages()[-1]
                shown = f"{avg:.2f} degC" if avg is not None else "no reading"
                print(f"  {now:%H:%M:%S}  avg {shown}", end="\r")
            time.sleep(interval)
        print(f"\nDone - {n} samples.")
    except KeyboardInterrupt:
        print(f"\nStopped early - {len(rec.rows)} samples kept.")
    finally:
        tc08.usb_tc08_close_unit(handle)

    if not rec.rows:
        raise SystemExit("No samples captured.")
    return rec


# ---------------------------------------------------------------- csv import

def load_csv(path):
    """Parse a PicoLog 6 CSV export: first column time, remaining columns temps."""
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = [r for r in reader if any(cell.strip() for cell in r)]
    if len(rows) < 2:
        raise SystemExit(f"{path}: not enough data rows.")

    header, data = rows[0], rows[1:]
    names = [h.strip() or f"CH{i}" for i, h in enumerate(header[1:], start=1)]
    rec = Recording(names)
    t0 = dt.datetime.now()

    for r in data:
        when = _parse_time(r[0], t0)
        temps = []
        for cell in r[1:len(names) + 1]:
            try:
                temps.append(float(cell.strip().replace(",", ".")))
            except ValueError:
                temps.append(None)
        rec.add(when, temps)

    if not rec.rows:
        raise SystemExit(f"{path}: no parsable rows.")
    return rec


def _parse_time(cell, t0):
    cell = cell.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S",
                "%m/%d/%Y %H:%M:%S", "%H:%M:%S"):
        try:
            parsed = dt.datetime.strptime(cell, fmt)
            if fmt == "%H:%M:%S":
                parsed = parsed.replace(year=t0.year, month=t0.month, day=t0.day)
            return parsed
        except ValueError:
            pass
    try:  # elapsed seconds since start
        return t0 + dt.timedelta(seconds=float(cell.replace(",", ".")))
    except ValueError:
        return t0


# ---------------------------------------------------------------- outputs

# palette: light chart surface + series blue + recessive chrome
SURFACE, SERIES, INK, MUTED, GRID, BASELINE = (
    "#fcfcfb", "#2a78d6", "#0b0b0b", "#898781", "#e1e0d9", "#c3c2b7")


def write_png(rec, path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.dates as mdates
    import matplotlib.pyplot as plt

    times, avgs = zip(*[(t, a) for t, a in zip(rec.times, rec.row_averages())
                        if a is not None])

    fig, ax = plt.subplots(figsize=(10, 5), facecolor=SURFACE)
    ax.set_facecolor(SURFACE)
    ax.plot(times, avgs, color=SERIES, linewidth=2)

    overall = rec.overall_average()
    ax.set_title(f"Average temperature - overall {overall:.2f} \N{DEGREE SIGN}C",
                 color=INK, fontsize=13, loc="left", pad=12)
    ax.set_ylabel("\N{DEGREE SIGN}C", color=MUTED)
    ax.grid(axis="y", color=GRID, linewidth=0.8)
    ax.tick_params(colors=MUTED)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color(BASELINE)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(path, dpi=150, facecolor=SURFACE)
    plt.close(fig)


def write_excel(rec, path):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Readings"
    ws.append(["Time"] + rec.channel_names + ["Average"])
    for cell in ws[1]:
        cell.font = bold
    for when, row, avg in zip(rec.times, rec.rows, rec.row_averages()):
        ws.append([when.strftime("%Y-%m-%d %H:%M:%S")]
                  + [round(t, 2) if t is not None else None for t in row]
                  + [round(avg, 2) if avg is not None else None])
    ws.column_dimensions["A"].width = 20

    n = len(rec.rows)
    avg_col = len(rec.channel_names) + 2
    chart = LineChart()
    chart.title = "Average temperature"
    chart.y_axis.title = "degC"
    chart.x_axis.title = "Time"
    chart.height, chart.width = 9, 18
    chart.add_data(Reference(ws, min_col=avg_col, min_row=1, max_row=n + 1),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    chart.series[0].graphicalProperties.line.solidFill = SERIES.lstrip("#")
    chart.series[0].graphicalProperties.line.width = 25000  # EMU ~2pt
    ws.add_chart(chart, f"{chr(ord('A') + avg_col + 1)}2")

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Channel", "Average (degC)", "Min", "Max", "Samples"])
    for cell in ws2[1]:
        cell.font = bold
    for name, mean, lo, hi, count in rec.channel_stats():
        ws2.append([name,
                    round(mean, 2) if mean is not None else "-",
                    round(lo, 2) if lo is not None else "-",
                    round(hi, 2) if hi is not None else "-",
                    count])
    overall = rec.overall_average()
    ws2.append([])
    ws2.append(["Overall average", round(overall, 2) if overall is not None else "-"])
    ws2["A" + str(ws2.max_row)].font = bold
    ws2.column_dimensions["A"].width = 18

    wb.save(path)


# ---------------------------------------------------------------- main

def main():
    parser = argparse.ArgumentParser(
        description="Graph + Excel of average temperature from a Pico TC-08 logger.")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_live = sub.add_parser("live", help="sample the TC-08 over USB")
    p_live.add_argument("--minutes", type=float, default=5.0, help="capture length (default 5)")
    p_live.add_argument("--interval", type=float, default=1.0, help="seconds between samples")
    p_live.add_argument("--channels", default="1,2,3,4,5,6,7,8",
                        help="comma list of channels 1-8 (default all)")
    p_live.add_argument("--tc-type", default="K", choices=list("BEJKNRST"),
                        help="thermocouple type (default K)")

    p_csv = sub.add_parser("csv", help="use a CSV exported from PicoLog 6")
    p_csv.add_argument("file", help="path to the exported CSV")

    for p in (p_live, p_csv):
        p.add_argument("--out-dir", default=".", help="output folder (default: current)")
        p.add_argument("--prefix", default=None, help="output file name without extension")

    args = parser.parse_args()

    if args.cmd == "live":
        channels = sorted({int(c) for c in args.channels.split(",") if c.strip()})
        if any(c < 1 or c > 8 for c in channels):
            raise SystemExit("Channels must be between 1 and 8.")
        rec = capture_live(args.minutes, args.interval, channels, args.tc_type)
        default_prefix = "picolog_avg_" + dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    else:
        rec = load_csv(args.file)
        default_prefix = os.path.splitext(os.path.basename(args.file))[0] + "_avg"

    prefix = args.prefix or default_prefix
    os.makedirs(args.out_dir, exist_ok=True)
    png = os.path.join(args.out_dir, prefix + ".png")
    xlsx = os.path.join(args.out_dir, prefix + ".xlsx")

    if all(a is None for a in rec.row_averages()):
        raise SystemExit("No valid temperature readings found - nothing to report.")

    write_png(rec, png)
    write_excel(rec, xlsx)

    overall = rec.overall_average()
    print(f"Overall average: {overall:.2f} degC over {len(rec.rows)} samples")
    print(f"Graph: {png}")
    print(f"Excel: {xlsx}")


if __name__ == "__main__":
    main()
